#!/usr/bin/python3
import openpyxl

wb = openpyxl.load_workbook("student.xlsx")
ws = wb['Sheet1']

row_id = 1;
score = []
for row in ws:
    if row_id != 1:
        sum_v = ws.cell(row = row_id, column = 3).value * 0.3
        sum_v += ws.cell(row = row_id, column = 4).value * 0.35
        sum_v += ws.cell(row = row_id, column = 5).value * 0.34
        sum_v += ws.cell(row = row_id, column = 6).value
        ws.cell(row = row_id, column = 7).value = sum_v
        score.append(sum_v)
    row_id += 1

#print(score)
studentNum = row_id - 2
sortedScore = sorted(score, reverse = True)
#print(sortedScore)
#print(studentNum)
a1 = int(studentNum * 0.3 * 0.5)
a0 = int(studentNum * 0.3 - a1)
b1 = int(studentNum * 0.4 * 0.5)
b0 = int(studentNum * 0.4 - b1)
c = studentNum - a1 - a0 - b1 - b0
#print(c)
c1 = int(c * 0.5)
c0 = int(c - c1)

#print(a1, a0, b1, b0, c1, c0)

#grade
grade = []
for s in range(1, studentNum + 1):
    if s <= a1:
        grade.append('A+')
    elif s <= (a1 + a0):
        grade.append('A0')
    elif s <= (a1 + a0 + b1):
        grade.append('B+')
    elif s <= (a1 + a0 + b1 + b0):
        grade.append('B0')
    elif s <= (a1 + a0 + b1 + b0 + c1):
        grade.append('C+')
    elif s <= (a1 + a0 + b1 + b0 + c1 + c0):
        grade.append('C0')

#print(grade)

row_id = 1 
for row in ws:
    if row_id != 1:
        index = 0
        for s in sortedScore:
            if ws.cell(row = row_id, column = 7).value == s:
                #print(s, index)
                break
            index += 1
        ws.cell(row = row_id, column = 8).value = grade[index]
        #print(grade[index])
    row_id += 1

wb.save("student.xlsx")
